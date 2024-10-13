import openpyxl
from openpyxl import Workbook

# Función para crear/cargar el archivo Excel
def cargar_o_crear_excel(archivo):
    try:
        return openpyxl.load_workbook(archivo)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        ws.append(["Fecha", "Descripción", "Monto"])
        wb.save(archivo)
        return wb
# Función para agregar un gasto
def agregar_gasto(ws, fecha, descripcion, monto):
    ws.append([fecha, descripcion, monto])

# Función para generar el resumen
def generar_resumen(ws):
    gastos = [(fila[0], fila[1], fila[2]) for fila in ws.iter_rows(min_row=2, values_only=True)]
    total = sum(g[2] for g in gastos)
    caro = max(gastos, key=lambda x: x[2])
    barato = min(gastos, key=lambda x: x[2])
    
    print(f"\nTotal gastos: {len(gastos)} | Gasto más caro: {caro[1]} ({caro[0]}) {caro[2]:.2f}")
    print(f"Gasto más barato: {barato[1]} ({barato[0]}) {barato[2]:.2f} | Monto total: {total:.2f}")
    
    ws.append([""])
    ws.append([f"Total de gastos: {len(gastos)}", f"Gasto más caro: {caro[1]} ({caro[0]}) {caro[2]:.2f}", f"Monto total: {total:.2f}"])

# Función principal para ingresar los datos
def ingresar_gastos():
    archivo = "informe_gastos.xlsx"
    wb = cargar_o_crear_excel(archivo)
    ws = wb["Gastos"]

    while True:
        fecha = input("Fecha del gasto (YYYY-MM-DD): ")
        descripcion = input("Descripción: ")
        try:
            monto = float(input("Monto: "))
        except ValueError:
            print("Monto inválido. Inténtalo de nuevo.")
            continue
        agregar_gasto(ws, fecha, descripcion, monto)

        if input("¿Agregar otro gasto? (s/n): ").lower() != 's':
            break

    generar_resumen(ws)
    wb.save(archivo)
    print("Informe guardado en 'informe_gastos.xlsx'.")

# Ejecutar el programa
if __name__ == "__main__":
    ingresar_gastos()
